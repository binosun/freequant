#!/user/bin/env python
#coding=utf-8

import pandas as pd

# 支持保存成xlsx和xls格式
def save_record(record, sheet_name, save_file_name):
    xlsx_output = pd.ExcelWriter(save_file_name)
    record.to_excel(xlsx_output, sheet_name=sheet_name)
    xlsx_output.save()



df = pd.read_excel('log.xls', sheetname=1,encoding="utf-8")

# 参数sheetname，表示的是第几个表，从0开始计数。我上面设置的是1，也就是第二个表。
# 当为字符串时，将查找名为字符串的sheet

# 新建一列，每一行的值是Member列和activity列相同行值的和
for i in df.index:
  df['activity_2'][i] = df['Member'][i] + df['activity'][i]

# 根据Member字段去除掉多余的行，并且保留相同行的最后一行数据
new_df = df.drop_duplicates(subset='Member', keep='last')


df=pd.read_csv('data/xxx.csv')
df.to_csv('result.csv', index=False, header=["year","state"]) #默认填写路径即可，如果不想添加索引（就是one two three这个），如果只想输出year和state这两个字段的数据，那么就如代码所示操作。
df.to_csv('/tmp/9.csv',columns=['open','high'],index=False,header=False)
# 不要列头，不要索引，只要open,high两列。



# 这篇文章主要介绍了python脚本实现xls(xlsx)转成csv的相关资料,需要的朋友可以参考下
# -*- coding: utf-8 -*-
import xlrd
import xlwt
import sys
from datetime import date, datetime


def read_excel(filename):
    workbook = xlrd.open_workbook(filename)
    sheet2 = workbook.sheet_by_index(0)

    for row in xrange(0, sheet2.nrows):
        rows = sheet2.row_values(row)

        def _tostr(cell):
            if type(u'') == type(cell):
                return "\"%s\"" % cell.encode('utf8')
            else:
                return "\"%s\"" % str(cell)

        print ','.join([_tostr(cell) for cell in rows])


if __name__ == '__main__':
    filename = sys.argv[1]
    read_excel(filename)

# 再给大家分享一则代码
#
# xlsx文件解析处理：openpyxl库
# csv文件格式生成：csv

# coding: utf-8
# 依赖openpyxl库：http://openpyxl.readthedocs.org/en/latest/

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import csv
import os
import sys


def xlsx2csv(filename):
    try:
        xlsx_file_reader = load_workbook(filename=filename)
        for sheet in xlsx_file_reader.get_sheet_names():
            # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
            csv_filename = '{xlsx}_{sheet}.csv'.format(
                xlsx=os.path.splitext(filename.replace(' ', '_'))[0],
                sheet=sheet.replace(' ', '_'))
            csv_file = file(csv_filename, 'wb')
            csv_file_writer = csv.writer(csv_file)
            sheet_ranges = xlsx_file_reader[sheet]
            for row in sheet_ranges.rows:
                row_container = []
                for cell in row:
                    if type(cell.value) == unicode:
                        row_container.append(cell.value.encode('utf-8'))
                    else:
                        row_container.append(str(cell.value))
                        csv_file_writer.writerow(row_container)
                        csv_file.close()
    except Exception as e:
        print(e)

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('usage: xlsx2csv <xlsx file name>')
else:
    xlsx2csv(sys.argv[1])
sys.exit(0)



from pyexcel.cookbook import merge_all_to_a_book
import pyexcel.ext.xlsx
# needed to support xlsx format, pip install pyexcel-xlsx
import glob

merge_all_to_a_book(glob.glob("your_csv_directory/*.csv"), "output.xlsx")